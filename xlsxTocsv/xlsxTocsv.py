#!/usr/bin/env python
# coding: utf-8
# datatab.py
# 从策划配置表目录 game\Design\配置表\”
#   生成服务器的 game\Program\server\six\datatab\” 目录，
#   和客户端的 game\Program\client\Assets\Config\” 目录。
#   所有xlsx文件生成csv文件，其他文件原样复制。
#   其中 server\ 目录特殊处理，仅对服务器有效，客户端跳过。
#
# 依赖openpyxl库：http://openpyxl.readthedocs.org/en/latest/
# 参考代码 http://segmentfault.com/q/1010000003006437?_ea=273128
# 测试环境：Python3.4

# Usage: datatab.py <game dir>
# Example: datatab.py "d:\game"
# <game dir> 是根目录，包含Design/, Program/ 目录。

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
import csv
import os
import sys
import shutil

def xlsx2csv(filename):
    # try:
        xlsx_file_reader = load_workbook(filename = filename, data_only = True)
        for sheet in xlsx_file_reader.get_sheet_names():
            # 仅第1个sheet输出到一个csv文件中，文件名后缀替换为.csv
            csv_filename = os.path.splitext(filename)[0] + '.csv'
            csv_file = open(csv_filename, 'w', encoding='utf8', newline='')
            csv_file_writer = csv.writer(csv_file)
            sheet_ranges = xlsx_file_reader[sheet]
            for row in sheet_ranges.rows:
                row_container = []
                for cell in row:
                    row_container.append(cell.value)
                csv_file_writer.writerow(row_container)
            # End of for row.
            csv_file.close()
            break  # 仅输出第1个sheet
        # End of for sheet.
    # End of try.
    # except Exception as e:
    #    print(e)
# End of xlsx2csv().

def datatab_convert(game_dir):
    '''从 game\Design\配置表\ 输出到
      game\Program\server\six\datatab\
      game\Program\client\Assets\Config\
    '''
    design_dir = os.path.join(game_dir, 'Design/配置表/')
    server_dir = os.path.join(game_dir, 'Program/server/six/datatab/')
    client_dir = os.path.join(game_dir, 'Program/client/Assets/Config/')

    # 删除旧文件。
    print("Delete " + server_dir)
    try:
        shutil.rmtree(server_dir)
    except:
        pass
    print("Delete " + client_dir)
    try:
        shutil.rmtree(client_dir)
    except:
        pass

    # 生成server文件
    print("Creating " + server_dir)
    shutil.copytree(design_dir, server_dir)
    files = get_files(server_dir)
    convert_files(files)

    # 复制client文件
    print("Copy " + client_dir)
    shutil.copytree(server_dir, client_dir)
    shutil.rmtree(os.path.join(client_dir, 'server/'))
    print("Done. Total files: %d" % len(files))
# End of datatab_convert().

def get_files(dir):
    '''Get a list of files under input dir.'''
    result = []
    for root,dirs,files in os.walk(dir):
        for f in files:
            result.append(os.path.join(root, f))
    return result
# End of get_files().

def convert_files(files):
    '''转换一批文件.
    files 是列表，元素为完整路径名。
    '''
    for f in files:
        ext = os.path.splitext(f)[1].lower()
        if '.xlsx' != ext:
            print(f + " -> keep")
            continue
        print(f + " -> csv")
        xlsx2csv(f)
        os.remove(f)
# End of convert_files().

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('usage: datatab <game dir>')
    else:
        datatab_convert(sys.argv[1])
    sys.exit(0)
